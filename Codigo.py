from openpyxl import Workbook
from openpyxl.styles import PatternFill

from PIL import Image
#from colors import rgb

im=Image.open('Teste.jpg')
pix=im.load()

wb=Workbook()
planilha=wb.worksheets[0]

for i in range(1,im.size[0]): 

    teste=planilha.cell(1,i).column_letter
    planilha.column_dimensions[teste].width=1

    for n in range(1,im.size[1]):
        
        planilha.row_dimensions[n].height=6

        r=str(hex(pix[i,n][0]))[2:]
        g=str(hex(pix[i,n][1]))[2:]
        b=str(hex(pix[i,n][2]))[2:]

        
        if len(r)<2:r="0"+r 
        if len(g)<2:g="0"+g 
        if len(b)<2:b="0"+b 

        #cor = Color(rgb=r+g+b)

        prenchimento = PatternFill(patternType='solid', fgColor=r+g+b)
        planilha.cell(n,i).fill=prenchimento

planilha.sheet_view.zoomScale=25      
print (str(im.size[0])+" x "+str(im.size[1]))
   


wb.save('Teste.xlsx')
